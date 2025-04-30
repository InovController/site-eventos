from django.shortcuts import get_object_or_404, render
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required
from .models import Participation
from events_manager.models import Event
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from django.utils import timezone


def export_participants_excel(request, event_id):
    event = Event.objects.get(id=event_id)
    participants = Participation.objects.filter(event=event).select_related('user')
    participants = sorted(participants, key=lambda p: not p.is_present)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Participantes"

    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center")

    # Informações do evento no topo
    ws["A1"], ws["B1"] = "Título do Evento:", event.title
    ws["A2"], ws["B2"] = "Local:", event.location
    ws["A3"], ws["B3"] = "Data:", event.date.strftime("%d/%m/%Y") if event.date else ""
    ws["A4"], ws["B4"] = "Horário:", event.time.strftime("%H:%M") if event.time else ""

    for row in range(1, 6):
        ws[f"A{row}"].font = bold_font

    start_row = 6
    headers = ["Nome", "E-mail", "Celular", "Empresa", "Checkin"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_num, value=header)
        cell.font = bold_font
        cell.alignment = center_alignment

    row_num = start_row + 1
    for p in participants:
        if p.is_present or p.signed_up:
            ws.cell(row=row_num, column=1, value=f"{p.user.first_name} {p.user.last_name}")
            ws.cell(row=row_num, column=2, value=p.user.email)
            ws.cell(row=row_num, column=3, value=p.user.phone)
            ws.cell(row=row_num, column=4, value=p.user.company)

            if p.date_joined:
                print('a')
                dt_local = timezone.localtime(p.date_joined)
                joined = dt_local.replace(tzinfo=None)
                checkin = joined.strftime("%d/%m/%Y %H:%M")
            else:
                checkin = "—"  # Ou "", se preferir vazio
            ws.cell(row=row_num, column=5, value=checkin)
            row_num += 1

    # Ajuste automático das colunas
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    # Gera resposta para download
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"participantes.xlsx".replace(" ", "_")
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response


def export_participants_pdf(request, event_id):
    event = Event.objects.get(id=event_id)
    participants = Participation.objects.filter(event=event).select_related('user')
    participants = sorted(participants, key=lambda p: not p.is_present)

    # Prepara a resposta
    response = HttpResponse(content_type='application/pdf')
    filename = f"participantes.pdf".replace(" ", "_")
    response['Content-Disposition'] = f'attachment; filename={filename}'

    # Criação do PDF
    doc = SimpleDocTemplate(response, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = []

    # Cabeçalho do evento
    elements.append(Paragraph(f"<b>{event.title.upper()}</b>", styles['Title']))
    elements.append(Spacer(1, 12))

    if event.location:
        elements.append(Paragraph(f"<b>Local:</b> {event.location}", styles['Normal']))
    if event.date:
        elements.append(Paragraph(f"<b>Data:</b> {event.date.strftime('%d/%m/%Y')}", styles['Normal']))
    if event.time:
        elements.append(Paragraph(f"<b>Horário:</b> {event.time.strftime('%H:%M')}", styles['Normal']))
    
    elements.append(Spacer(1, 18))

    # Cabeçalhos da tabela
    data = [["Participante", "Contato", "Data/Hora Check-in"]]

    # Participantes (presentes primeiro)
    participantes_ordenados = sorted(participants, key=lambda p: not p.is_present)

    for p in participantes_ordenados:
        if p.is_present or p.signed_up:
            nome = f"{p.user.first_name} {p.user.last_name}\n{p.user.company}"
            email = p.user.email or ""
            phone = p.user.phone or ""
            contato = f"{email}\n{phone}".strip()
            if p.date_joined:
                dt_local = timezone.localtime(p.date_joined)
                checkin = dt_local.strftime("%d/%m/%Y %H:%M")
            else:
                checkin = "—"  # Ou "", se preferir vazio
            data.append([nome, contato, checkin])

    # Cria tabela com estilo
    table = Table(data, colWidths=[180, 180, 130])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('FONTSIZE', (0,0), (-1,-1), 10),
        ('BOTTOMPADDING', (0,0), (-1,-1), 6),
        ('GRID', (0,0), (-1,-1), 0.25, colors.grey),
    ]))

    elements.append(table)
    doc.build(elements)
    return response


@login_required(login_url='login')
def subscribe_event(request, pk):
    if request.method == 'POST':
        event = get_object_or_404(Event, id=pk)
        Participation.objects.create(event=event, user=request.user, signed_up=True)
        return JsonResponse({'success': True})

    return JsonResponse({'success': False, 'error': 'Invalid request'}, status=400)


@login_required(login_url='login')
def unsubscribe_event(request, pk):
    if request.method == 'POST':
        event = get_object_or_404(Event, id=pk)
        Participation.objects.filter(event=event, user=request.user).delete()
        return JsonResponse({'success': True})

    return JsonResponse({'success': False, 'error': 'Invalid request'}, status=400)


@login_required(login_url='login')
def validate_presence(request, token):
    event = get_object_or_404(Event, token=token)
    
    if event.status == 'encerrado':
        return render(request, 'no_validate_participation.html', {
            'object': event,
            'message': 'Desculpe, este evento já foi encerrado.',
        })
        

    participation, created = Participation.objects.get_or_create(event=event, user=request.user)

    if participation.is_present:
        return render(request, 'validate_participation.html', {
            'object': event,
            'message': 'Você já registrou sua presença para o evento',  # Adiciona uma mensagem informativa
            'participation': participation,
        })

    participation.is_present = True
    participation.date_joined = timezone.now()
    participation.save()

    return render(request, 'validate_participation.html', {
        'object': event,
        'message': 'Sua presença foi registrada com sucesso para o evento',  # Mensagem de sucesso
        'participation': participation,
    })
